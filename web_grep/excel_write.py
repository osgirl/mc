from openpyxl import Workbook
import xlsxwriter
#import XlsxWriter
import datetime
def create_output():
    wb = Workbook()
    ws1 = wb.active
    #xbook = xlsxWriter.Workbook('Test.xlsx')
    #xbook = x
    #xsheet = xbook.add_worksheet('Test')
    #a=['123','outof']
    #xsheet.write(0,0,''.join(a))
    #xbook.close()
    ws1 = wb.create_sheet(datetime.date.today().strftime("%Y-%m-%d"),0)
    ws1.cell(row=1, column=1, value='Product ID')
    ws1.cell(row=1, column=2, value='Availablity')
    ws1.cell(row=1, column=3, value='Sale')
    ws1.cell(row=1, column=4, value='Size')
    wb.save('output.xlsx')
    #a='123'+'\n'+'outof'+'\n'+'test'

    #ws1.cell(row=1,column=5,value=a)
    #ws1.cell(row=2,column=5,value=a)
    #wb.save('output.xlsx')
    return wb
#create_output()

