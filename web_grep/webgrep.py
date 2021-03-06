import re
import urllib.request
import urllib.parse
import excel_read
from openpyxl import Workbook
import argparse
import datetime
import os
size=['tiny baby','new baby','Up to 1 mnth','Up to 3 mnths','0-6 months','6-12 months','1-2 years','2-4 years','4-7 years','3-6 months','6-9 months','9-12 months','12-18 months','18-24 months','2-3 years','3-4 years','4-5 years','5-6 years','6-7 years','7-8 years','1 adlt','2 adlt','4 jnr','5 jnr','6 jnr','7 jnr','8 jnr','9 jnr','10 jnr','11 jnr','12 jnr','13 jnr']
size1=['0-6 months','6-12 months','1-2 years','2-4 years','4-7 years','3-6 months','6-9 months','9-12 months','12-18 months','18-24 months','2-3 years','3-4 years','4-5 years','5-6 years','6-7 years','7-8 years','1 adlt','2 adlt','4 jnr','5 jnr','6 jnr','7 jnr','8 jnr','9 jnr','10 jnr','11 jnr','12 jnr','13 jnr']
url = 'https://www.mothercare.com/search'
#productid=['PJ143']
import excel_write


def salecheck(stock_lists):
    ava1=0
    for sale_item in stock_lists:
            if 'in stock' in sale_item :
                for size_list in size1:
                    if size_list in sale_item:
                         ava1=1
                        # print('ava1=1')
                         break              
    #             else:
    #                ava1=0
    return ava1

if __name__=='__main__':
    parser = argparse.ArgumentParser(description='webgrep for MC product')
    #parser.add_argument('input_file', default='/home/yu/Desktop/product_test.xlsx',
    #               help='input file path') 
    #parser.add_argument('input_file', nargs='?', default='product_test.xlsx',help='input file path')
    #args = parser.parse_args()
    filename=os.getenv('FILE_NAME', '/home/yu/Mothercare/web_grep/product_total.xlsx')
    productid=excel_read.product_check(filename)
    outputwb=excel_write.create_output()
    output=outputwb[datetime.date.today().strftime("%Y-%m-%d")]
    sale_counter=0
    no_availablity_counter=0
    for i in range(0,len(productid)):
        print('Checking Product '+productid[i])
        try:
            values = {'q':productid[i]}
            data = urllib.parse.urlencode(values)
            data = data.encode('utf-8')
            req = urllib.request.Request(url, data)
            resp = urllib.request.urlopen(req)
            respData = resp.read().decode('utf-8')
            #f=open('temp1','w')
            #f.write(respData)
            #f.close()
            result=[]
            stock_lists=[]
            ava=1
            salecount=0
            line = respData.splitlines()
            for i in range (1,len(line)):
                if 'no products were found for your search:' in line[i]:
                    ava=0
                    break
                if 'sale' in line[i] and len(line[i])<5:
                    salecount=salecount+1
                if any(word in line[i] for word in size):
                   # siset(line[i])
                   # sizetmp=set(line[i])-(set(line[i])-set(size))
                   # print(sizetmp)
                   product_size="none"
                   for sizetmp in size:
                        if sizetmp in line[i]:
                            product_size=sizetmp
                            break

                   if 'data-enabled' in line[i-1]:
                       # print(line[i])
                       # print(line[i-1])
                        if 'true' in line[i-1]:
                            stock_list=str(product_size+' in stock\n')
                        else:
                            stock_list=str(product_size+' out of stock')
                        stock_lists.append(stock_list+'\n')
                   # print(stock_lists)
            if ava==0:
                no_availablity_counter=no_availablity_counter+1
                row_max=output.max_row
                output.cell(row=row_max+1, column=1, value=values['q'])
                output.cell(row=row_max+1, column=2, value='N')
                #print('==============================================')
                #print('Product Code: '+values['q']+'\n')
                #print('No product available')
            if salecount>1 and ava==1:
                sale_counter=sale_counter+1
                #print(stock_lists)
                ava=salecheck(stock_lists)
        #       print(ava)
                #for sale_item in stock_lists:
                #    if 'in stock' in sale_item :
                #        for size_list in size1:
                #            if size_list in sale_item:
                #                ava=1
                #                break
                #    else:
                #        ava=0
                #        continue
                #    break
                if ava==1:
                    row_max=output.max_row
                    output.cell(row=row_max+1, column=1, value=values['q'])
                    output.cell(row=row_max+1, column=2, value='Y')
                    output.cell(row=row_max+1, column=3, value='Y')
                    output.cell(row=row_max+1, column=4, value=str.join('.\n', stock_lists))
                if ava==0:
                    no_availablity_counter=no_availablity_counter+1
                    row_max=output.max_row
                    output.cell(row=row_max+1, column=1, value=values['q'])
                    output.cell(row=row_max+1, column=2, value='N')
                    output.cell(row=row_max+1, column=3, value='Y')
                    output.cell(row=row_max+1, column=4, value=str.join('.\n', stock_lists))

                #print('==============================================')
                #print('Product Code: '+values['q']+'\n')
                #print('Product On Sale')
                #for i in range(1,len(stock_lists)):
                #    print(stock_lists[i]+'\n')
        except  urllib.error.URLError as e:
            print(productid[i]+'\n')
            print(e.reason)
    outputwb.save('Mothercare'+datetime.date.today().strftime("%Y-%m-%d")+'.xlsx')
    try:
        outputwb.save('/app/mc/'+'Mothercare'+datetime.date.today().strftime("%Y-%m-%d")+'output.xlsx')
    except FileNotFoundError as e:
        print("Skip file save in /app/mc")
    
    print('sale count = ',sale_counter)
    print('no availablity count = ', no_availablity_counter)
    print('Job Done')






