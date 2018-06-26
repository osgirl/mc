#!/usr/bin/env /usr/local/bin/python3
from openpyxl import load_workbook
#from xlrd import open_workbook
import re
import warnings
import os.path
import os
from time import gmtime, strftime

warnings.filterwarnings('ignore')


import urllib.request
import urllib.parse
url = 'https://www.mothercare.com/search'
class Product:
    def __init__(self,id,size):

        self.id=id;
        self.size=size;
    def check(self):
        ava=1;
        try:
            print('Checking Product '+self.id)
            values = {'q':self.id}
            data = urllib.parse.urlencode(values)
            data = data.encode('utf-8')
            req = urllib.request.Request(url, data)
            resp = urllib.request.urlopen(req)
            respData = resp.read().decode('utf-8')
            line = respData.splitlines()
            #print(line)
            #print(self.size)
            for i in range (1,len(line)):
                #print(line[i])
                if 'no products were found for your search:' in line[i]:
                    ava=3
                    break
                if self.size!='Any' and self.size in line[i] and 'out of stock' not in line[i] and len(line[i])<30:
                    ava=1
                    print(line[i])
                    break
                    #return 1
                if self.size!='Any' and self.size in line[i] and 'out of stock' in line[i] and len(line[i])<30:
                    ava=0
                    #print('out of stock')
                    print(line[i])
                    break
                else:
                    ava=1
        except urllib.error.URLError as e:
            print(self.id+'\n')
            print(e.reason)
            ava=2;
        return ava
def read_file(filename):
    wb = load_workbook(filename)
    wb1 = wb['Sheet1']
    itemsids=[]
    itemsizes=[]
    #print(wb1.max_row)
    for row in range (1,wb1.max_row):
        cell1=wb1.cell(row=row+1, column=1)
        itemid=cell1.value
        #itemid= wb1.cell(row=row+1, column=1)
        itemsids.append(itemid)
        cell2=wb1.cell(row=row+1, column=2)
        #itemsize=wb1.cell(row=row+1,column=2).
        itemsize=cell2.value
        itemsizes.append(itemsize)
        products=[itemsids,itemsizes]
    return products



if __name__ == "__main__":
    #Test=product('MA560','any')
    #result=Test.check()
    #test=read_file('test.xlsx')
    #print(test)
    products=read_file('product_check_list.xlsx')
    #print(products)
    cmd1='cat temp.txt | mail -s Product_Available_Now pmc_tina@hotmail.com'
    cmd2='rm temp.txt'
    ava_results=[]
    ava_id=[]
    ava_size=[]
    for i in range (0,len(products[0])):
        product=[]
        product=Product(products[0][i],products[1][i])
        check_result=product.check()
        ava_id.append(products[0][i])
        ava_size.append(products[1][i])
        ava_results.append(check_result)
    for i in range (0,len(ava_results)):
        if ava_results[i]==1:
            with open('temp.txt','w+') as f:
                f.write('Product ID: '+ava_id[i]+'.  Product Size: '+ava_size[i])
                f.write('\n Checking Time: '+strftime("%Y-%m-%d %H:%M:%S", gmtime()))
                f.close()
            print('Product Available')
    if os.path.isfile('temp.txt'):
        try:
            os.system(cmd1)
            os.system(cmd2)
            print('Email Send')
        except OSError as err:
            print(err)





